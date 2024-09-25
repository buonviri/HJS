import openpyxl
from openpyxl.styles import Font, Alignment

# set to false to skip formatting and sample data
formatted = True

wb = openpyxl.Workbook()
if formatted:
    ws = wb.active
    info = [[],['','a','b','c'],['',1,2,3],['',4,5,6]]  # has blank row and col
    for row in info:
        ws.append(row)  # add data to worksheet
    bold = Font(bold=True)  # create bold font
    centered = Alignment(horizontal='center', vertical='center')  # create centered alignment
    for row in ws['B2:D2']:
        for cell in row:
            cell.font = bold  # set each cell to bold
    for row in ws['B2:D4']:
        for cell in row:
            cell.alignment = centered  # center each cell

#save
wb.save('new.xlsx')
