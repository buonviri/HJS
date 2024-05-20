# instructions:
# run this file in a folder containing one or more xlsx files
# each will be converted to tab data and stored as a new file
# the last file read will be placed on the clipboard as well

import os
import pyperclip
from openpyxl import load_workbook

file_list = []
extensions = ('.xlsm', '.xlsx')  # must be five chars long
for file in os.listdir('.'):
    if file[-5:] in extensions:
        file_list.append(file)

clip = ''  # blank string to store clipboard info
for x in file_list:
    print()
    print('  Reading: ' + x)

    wb = load_workbook(filename=x, read_only=True)  # load as read-only
    ws = wb[wb.sheetnames[0]]  # select the first sheet

    found_value = False
    orcad_bom = False
    for i in ws.values:
        row = []  # blank string to hold row
        for j in i:
            if j is None:
                j = ''
            if len(j) > 0 and not found_value:
                found_value = True
                if j == 'Table':  # special case for orcad BOMs, could add additional checks if needed
                    orcad_bom = True
                    print('    Found "Table", setting orcad_bom to True')
            row.append(j.strip())  # remove unwanted whitespace
        if orcad_bom:  # modify some columns
            row[0] = row[0].rjust(30, '.')
            row[1] = row[1].rjust(30, '.')
        clip = clip + '\t'.join(row) + '\n'

    newfile = x[:-5] + '.tab'
    with open(newfile, "w") as f:
        f.write(clip)
    print('  Writing: ' + newfile)

    # Close the workbook after reading
    wb.close()
    del wb

# copy to clipboard and exit
pyperclip.copy(clip)
print('  Copying: ' + newfile[:-4])  # remove tab extension
print()
os.system('PAUSE')
